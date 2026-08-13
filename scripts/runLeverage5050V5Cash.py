#!/usr/bin/env python3
from __future__ import annotations

import json
import sys
import urllib.request
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from leverage5050Backtest import ETF_COSTS, StrategySpec, align_market, benchmark_buy_hold, index_benchmark
from leverage5050CashBacktest import simulate_with_cash_curve
from scripts.runLeverage5050Backtest import START_FLOOR, SPLITS, adjust_split, fetch_stock, fetch_taiwan50, month_keys
from scripts.runLeverage5050V2 import WINDOWS, strict_gate

MONTH_COUNT = 144
CBC_URL = "https://www.cbc.gov.tw/tw/public/data/a13rate.xls"
BANK_SHEETS = ["台銀", "合庫", "土銀", "華銀", "一銀"]
# Workbook row layout: col 0 ROC YYMM; zero-based col 13 = 1-year fixed term-deposit rate.
ONE_YEAR_FIXED_COL = 13
PRIMARY = "asym35_77_5_t50"
CASH_FACTORS = [0.50, 0.75, 1.00]


def label(x: float) -> str:
    p = x * 100
    if abs(p - round(p)) < 1e-9:
        return str(int(round(p)))
    return f"{p:.1f}".replace(".", "_").rstrip("0").rstrip("_")


def band_signal(lower: float, upper: float, target: float):
    def signal(i, rows, weight, state):
        if i == 0:
            return target
        return target if weight < lower or weight > upper else None
    return signal


def local_specs():
    # Pre-registered before v5 results: a tight neighborhood around the v3 near-pass.
    lows = [0.34, 0.35, 0.36]
    highs = [0.765, 0.775, 0.785]
    targets = [0.49, 0.50, 0.51]
    out = []
    for low in lows:
        for high in highs:
            for target in targets:
                name = f"asym{label(low)}_{label(high)}_t{label(target)}"
                out.append(StrategySpec(name, "local_asymmetric_band", band_signal(low, high, target)))
    assert len(out) == 27
    assert PRIMARY in {s.name for s in out}
    return out


def fetch_cbc_rates():
    req = urllib.request.Request(CBC_URL, headers={"User-Agent": "Mozilla/5.0 STOCK-leverage-research/1.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = resp.read()
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    bank_rates = {}
    for sheet in BANK_SHEETS:
        ws = wb[sheet]
        rates = {}
        for values in ws.iter_rows(min_row=6, values_only=True):
            code = values[0]
            rate = values[ONE_YEAR_FIXED_COL] if len(values) > ONE_YEAR_FIXED_COL else None
            if code is None or rate is None:
                continue
            try:
                code_s = str(int(float(code))).zfill(5)
                roc_year = int(code_s[:-2])
                month = int(code_s[-2:])
                if not 1 <= month <= 12:
                    continue
                year = roc_year + 1911
                key = f"{year:04d}-{month:02d}"
                rates[key] = float(rate) / 100.0
            except Exception:
                continue
        bank_rates[sheet] = rates

    months = sorted(set().union(*(set(v) for v in bank_rates.values())))
    avg = {}
    for month in months:
        vals = [bank_rates[s].get(month) for s in BANK_SHEETS]
        vals = [v for v in vals if v is not None]
        if len(vals) >= 4:
            avg[month] = sum(vals) / len(vals)
    if not avg:
        raise RuntimeError("CBC deposit-rate parser produced no monthly rates")
    return avg, bank_rates


def previous_month(date_str: str) -> str:
    y, m = int(date_str[:4]), int(date_str[5:7])
    m -= 1
    if m == 0:
        y -= 1
        m = 12
    return f"{y:04d}-{m:02d}"


def make_cash_curve(avg_rates, factor: float):
    keys = sorted(avg_rates)
    def annual(date_str: str) -> float:
        target = previous_month(date_str)  # one-month information lag
        if target in avg_rates:
            return avg_rates[target] * factor
        eligible = [k for k in keys if k <= target]
        return avg_rates[eligible[-1]] * factor if eligible else 0.0
    return annual


def compact(r):
    return {"name": r["name"], "family": r.get("family"), "period": r["period"], "metrics": r["metrics"]}


def run_window(market, tai50, specs, cash_curve, start, end):
    a = max(start, market[0]["date"])
    z = min(end, market[-1]["date"])
    window_market = [r for r in market if a <= r["date"] <= z]
    common_dates = {r["date"] for r in window_market}
    aligned_tr = [r for r in tai50 if r["date"] in common_dates]
    if len(aligned_tr) != len(window_market):
        raise RuntimeError(f"calendar mismatch {a}..{z}: market={len(window_market)} tr={len(aligned_tr)}")
    b0050 = benchmark_buy_hold(market, start_date=a, end_date=z)
    btr = index_benchmark(aligned_tr, start_date=a, end_date=z, value_key="taiwan50TR")
    results = [simulate_with_cash_curve(
        market, s, start_date=a, end_date=z, annual_cash_yield=cash_curve
    ) for s in specs]
    return a, z, b0050, btr, results


def evaluate_factor(market, tai50, specs, avg_rates, factor):
    windows = {"full": (market[0]["date"], market[-1]["date"]), **WINDOWS}
    benchmarks, per, wr = {}, {}, {}
    curve = make_cash_curve(avg_rates, factor)
    for key, (ws, we) in windows.items():
        a, z, b0050, btr, results = run_window(market, tai50, specs, curve, ws, we)
        benchmarks[key] = {"0050": b0050, "taiwan50TotalReturn": btr}
        wr[key] = {"period": {"start": a, "end": z}, "benchmarks": benchmarks[key]}
        for r in results:
            per.setdefault(r["name"], {"name": r["name"], "family": r["family"]})[key] = compact(r)
    ranking = []
    for block in per.values():
        checks, passed, wins = strict_gate(block, benchmarks)
        block["gate"] = checks
        block["gatePassed"] = passed
        block["regimeWins"] = wins
        ranking.append(block)
    ranking.sort(key=lambda r: (
        r["gatePassed"], r["regimeWins"],
        r["full"]["metrics"]["sharpe"], r["full"]["metrics"]["totalReturn"]
    ), reverse=True)
    return {"factor": factor, "windows": wr, "ranking": ranking,
            "passingCount": sum(1 for r in ranking if r["gatePassed"])}


def pct(x): return "—" if x is None else f"{x * 100:.2f}%"
def num(x): return "—" if x is None else f"{x:.4f}"


def markdown(report):
    f50 = report["factors"]["0.50"]
    benchmark = f50["windows"]["full"]["benchmarks"]["taiwan50TotalReturn"]
    lines = [
        "# Leverage 50/50 Challenger v5 — Official CBC Cash Yield",
        "",
        f"Period: **{report['period']['start']} → {report['period']['end']}**",
        "",
        "No signal parameter from prior rounds is rescued ex post. The pre-registered primary is `asym35_77_5_t50` (rebalance 00631L back to 50% only when its portfolio weight falls below 35% or rises above 77.5%). v5 evaluates a 3×3×3 local neighborhood and applies the official five-bank average 1-year fixed term-deposit rate with a one-month information lag.",
        "",
        "Cash-rate robustness: 50%, 75%, and 100% of the lagged official rate are tested. The 50% haircut is the acceptance case, to reflect that a fully liquid cash sleeve may not continuously earn the headline one-year fixed-deposit rate.",
        "",
        f"Aligned Taiwan 50 Total Return benchmark: {pct(benchmark['totalReturn'])}, DD {pct(benchmark['maxDrawdown'])}, Sharpe {num(benchmark['sharpe'])}.",
        "",
        "## Primary candidate by cash-yield assumption",
        "",
        "| CBC rate factor | Return | Max DD | Sharpe | Holdout return | Holdout Sharpe | Regime wins | Gate | Passing neighbors |",
        "|---:|---:|---:|---:|---:|---:|---:|---|---:|",
    ]
    for key in ["0.50", "0.75", "1.00"]:
        block = report["factors"][key]
        primary = next(r for r in block["ranking"] if r["name"] == PRIMARY)
        m, h = primary["full"]["metrics"], primary["holdout"]["metrics"]
        lines.append(
            f"| {float(key):.0%} | {pct(m['totalReturn'])} | {pct(m['maxDrawdown'])} | {num(m['sharpe'])} | "
            f"{pct(h['totalReturn'])} | {num(h['sharpe'])} | {primary['regimeWins']}/3 | "
            f"{'PASS' if primary['gatePassed'] else '—'} | {block['passingCount']}/27 |"
        )
    lines += ["", "## Acceptance verdict", ""]
    if report["robustMethodFound"]:
        lines += [
            "**Robust method found under the pre-registered gate.**",
            "",
            f"The primary passes even with only 50% of the official lagged deposit rate, and {report['factors']['0.50']['passingCount']} / 27 local neighbors pass the same strict Total Return gate. The primary also passes at 75% and 100% cash-rate assumptions.",
        ]
    else:
        lines += [
            "**No robust method yet.** The primary must pass at all three cash-rate factors and at least five of 27 local neighbors must pass under the 50% haircut case.",
        ]
    lines += [
        "",
        f"CBC rate coverage used: {report['cashRateCoverage']['start']} → {report['cashRateCoverage']['end']}; mean annual five-bank 1Y fixed rate over available months {pct(report['cashRateCoverage']['mean'])}.",
        "",
        "External ETF costs and split adjustments are unchanged. Cash interest is modeled pre-tax. Research only, not investment advice.",
    ]
    return "\n".join(lines) + "\n"


def main():
    months = month_keys(MONTH_COUNT)
    with ThreadPoolExecutor(max_workers=4) as pool:
        f0050 = pool.submit(fetch_stock, "0050", months)
        f631 = pool.submit(fetch_stock, "00631L", months)
        ft50 = pool.submit(fetch_taiwan50, months)
        fcash = pool.submit(fetch_cbc_rates)
        rows0050 = adjust_split(f0050.result(), "0050")
        rows631 = adjust_split(f631.result(), "00631L")
        tai50 = ft50.result()
        avg_rates, bank_rates = fcash.result()

    market = [r for r in align_market(rows631, rows0050) if r["date"] >= START_FLOOR]
    if len(market) < 2500:
        raise RuntimeError(f"insufficient common history: {len(market)}")
    specs = local_specs()
    factors = {}
    for factor in CASH_FACTORS:
        factors[f"{factor:.2f}"] = evaluate_factor(market, tai50, specs, avg_rates, factor)

    primary_pass_all = all(
        next(r for r in factors[k]["ranking"] if r["name"] == PRIMARY)["gatePassed"]
        for k in ["0.50", "0.75", "1.00"]
    )
    robust = bool(primary_pass_all and factors["0.50"]["passingCount"] >= 5)
    rate_values = [avg_rates[k] for k in sorted(avg_rates)]
    report = {
        "version": "leverage-5050-v5-cbc-cash",
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "period": {"start": market[0]["date"], "end": market[-1]["date"]},
        "primaryCandidate": PRIMARY,
        "cashRateSource": "CBC five leading banks, 1-year fixed term-deposit rates; monthly bank average; one-month lag",
        "cashRateFactors": CASH_FACTORS,
        "cashRateCoverage": {"start": min(avg_rates), "end": max(avg_rates), "mean": sum(rate_values) / len(rate_values)},
        "costs": ETF_COSTS,
        "corporateActions": SPLITS,
        "factors": factors,
        "robustMethodFound": robust,
    }
    out = ROOT / "data/backtests/leverage_5050_v5.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2, allow_nan=False) + "\n", encoding="utf-8")
    doc = ROOT / "docs/LEVERAGE_5050_V5.md"
    doc.parent.mkdir(parents=True, exist_ok=True)
    doc.write_text(markdown(report), encoding="utf-8")
    print(doc.read_text(encoding="utf-8"))


if __name__ == "__main__":
    main()
